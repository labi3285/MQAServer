import os
from django.http.response import HttpResponse
from django.http import StreamingHttpResponse
from django.http import JsonResponse
from aws_mqaserver.utils import value
from django import forms
import getpass
import openpyxl
import warnings
warnings.filterwarnings("ignore")
import psycopg2
from sqlalchemy import create_engine
import urllib
import json
import pandas as pd
from wsgiref.util import FileWrapper
from aws_mqaserver.models import *

global levelListPath #层级列表地址
global rootFolderPath #根目录地址
global checkListPath #根据产线名拼接的checklist地址
global lineFolderPath #当前产线根目录
global visitHistoryPath #历史visit文件夹地址
global productLineDict #五级产线字典


# 获取产线层级列表
def downloadLevelContent(request):
    level_list_path = './siteLevelData.json'
    local_file = open(level_list_path, 'rb')
    json_file = json.load(local_file)

    if request.method == 'GET':
        response = HttpResponse(value.ResponseData(json_file), \
                                content_type='application/octet-stream; charset=utf-8')
        return response
    else:
        return HttpResponse('error')


def file_name(file_dir):
    L=[]
    for root,dirs,files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1]=='.json':
                L.append(os.path.join(root, file))
    return (L)


# 下载checklist
def downloadCheckList(request):
    lob = request.POST.get('lob')
    site = request.POST.get('site')
    productLine = request.POST.get('product_line')
    project = request.POST.get('project')
    part = request.POST.get('part')
    checkListPath = rootFolderPath + lob + site + productLine + project + part + '/Checklist/'
    lineFolderPath = rootFolderPath + lob + site + productLine + project + part

    check_empty = file_name(checkListPath)
    if len(check_empty) == 0:
        error = 'CheckList not exists'
        # print(checkListPath)
        return HttpResponse(error)
    else:
        if request.method == 'POST':
            checklistName = lob + '_' + site + '_' + productLine + '_' + project + '_' + part + '.json'
            checkListFilePath = checkListPath + checklistName
            wrapper = FileWrapper(open(checkListFilePath, 'rb'))
            response = StreamingHttpResponse(wrapper)
            response['Content-Type'] = 'application/octet-stream'
            response['Content-Disposition'] = 'attachment;filename="{}"'.format(checkListPath)
            return response


# 获取历史visit
def donwnloadVisitHistory(request):
    visitHistoryPath = lineFolderPath + '/History/Visit/'
    if not os.listdir(visitHistoryPath):
        return HttpResponse('CheckList not exists')
    else:
        if request.method == 'GET':
            wrapper = FileWrapper(open(visitHistoryPath, 'rb'))
            response = StreamingHttpResponse(wrapper)
            response['Content-Type'] = 'application/octet-stream'
            response['Content-Disposition'] = 'attachment;filename="{}"'.format(checkListPath)
            return response


class UserForm(forms.Form):
    filename = forms.FileField()


# 上传visit
def uploadVisit(request):
    if request.method == "POST":
        userfrom = UserForm(request.POST, request.FILES)
        if userfrom.is_valid():
            lob = request.POST.get('lob')
            site = request.POST.get('site')
            productLine = request.POST.get('product_line')
            project = request.POST.get('project')
            part = request.POST.get('part')
            file_save_path = str(lob) + str(site) + str(productLine) + str(project) + str(part) + '/'

            filename = userfrom.cleaned_data['filename']
            f = Upload()
            f.fullpath = 'static/' + file_save_path + 'History/Visit/'
            f.file_path = filename
            f.save()
            return JsonResponse({"result": 200, "msg": "请求成功, 上传文件"})

        return JsonResponse({"result": 201, "msg": "请求失败"})
    else:
        return JsonResponse({"result": 202, "msg": "请求失败，请求方式错误"})


class FileFieldForm(forms.Form):
    filetype = forms.CharField()
    file_field = forms.FileField(widget=forms.ClearableFileInput(attrs={'multiple': True}))
    # filename = forms.FileField()


def upload_image_video(request):
    if request.method == 'POST':
        fileForm = FileFieldForm(request.POST, request.FILES)
        # userfrom = FileFieldForm(request.POST, request.FILES)
        if fileForm.is_valid():
            filetype = fileForm.cleaned_data['filetype']
            files = request.FILES.getlist('file_field')
        #     filename = ''
        #     for f in files:
        #         print(f)
        #     file = uploadImgandVideo()
            # file.fullpath = 'AcousticAACiPadX1751Woofer-Tweeter/' + 'History/'
            # file.file_path = filename
            # file.save()
            return JsonResponse({"result": 200, "msg": "请求成功, 上传文件"})

        return JsonResponse({"result": 201, "msg": "请求失败"})
    else:
        return JsonResponse({"result": 202, "msg": "请求失败，请求方式错误"})


def upload_record_excel(request):
    if request.method == "POST":
        userfrom = UserForm(request.POST, request.FILES)
        # print(request.FILES)
        if userfrom.is_valid():
            filename = userfrom.cleaned_data['filename']
            lob = request.POST.get('lob')
            site = request.POST.get('site')
            productLine = request.POST.get('product_line')
            project = request.POST.get('project')
            part = request.POST.get('part')
            file_save_path = str(lob) + str(site) + str(productLine) + str(project) + str(part) + '/'

            f = UploadRecord()
            #f.savepath = '/Users/mqa_server/ServerDataBase/Acoustic/AcousticAACiPadX1751Woofer-Tweeter/' + 'History/Record/'
            f.fullpath = 'static/' + file_save_path + 'History/Record/'
            f.file_path = filename
            f.save()
            return JsonResponse({"result": 200, "msg": "请求成功, 上传record"})

        return JsonResponse({"result": 201, "msg": "请求失败"})
    else:
        return JsonResponse({"result": 202, "msg": "请求失败，请求方式错误"})


def upload_mil_excel(request):
    if request.method == "POST":
        userfrom = UserForm(request.POST, request.FILES)
        if userfrom.is_valid():
            filename = userfrom.cleaned_data['filename']
            lob = request.POST.get('lob')
            site = request.POST.get('site')
            product_line = request.POST.get('product_line')
            project = request.POST.get('project')
            part = request.POST.get('part')
            file_save_path = str(lob) + str(site) + str(product_line) + str(project) + str(part) + '/'

            f = UploadMIL()
            f.fullpath = 'static/' + file_save_path + 'History/MIL/'
            f.file_path = filename
            f.save()
            return JsonResponse({"result": 200, "msg": "请求成功, 上传MIL"})

        return JsonResponse({"result": 201, "msg": "请求失败"})
    else:
        return JsonResponse({"result": 202, "msg": "请求失败，请求方式错误"})


def get_by_project_score(df):
    def cal_score(score_list):
        initial = 100
        cri_amount = score_list.count('Critical')
        major_amount = score_list.count('Major')
        minor_amount = score_list.count('Minor')
        final_score = initial - cri_amount * 20 - major_amount * 5 - minor_amount * 2
        return (final_score)

    def cal_range(score):
        if score >= 90:
            return (1)
        elif score >= 85:
            return (2)
        else:
            return (3)

    result = pd.DataFrame(columns=['Vendor', 'Year', 'Month', 'Project', 'Audit type', 'Score', 'Range'])

    df_temp = df.copy(deep=True)
    df_temp['unique'] = '' * df_temp.shape[0]
    for i in range(df_temp.shape[0]):
        df_temp['unique'][i] = str(df_temp['vendor'][i]) + str(df_temp['Year'][i]) + str(df_temp['Month'][i]) + str(
            df_temp['Project'][i]) + str(df_temp['Audit type'][i])

    unique_set = set(df_temp['unique'].tolist())
    for i in unique_set:
        df_record = df_temp[df_temp['unique'] == i]
        score_list = df_record['Issue sevrity'].tolist()
        score = cal_score(score_list)
        score_range = cal_range(score)
        temp_loc = result.shape[0]
        result.loc[temp_loc + 1] = [str(df_record['vendor'].tolist()[0]), str(df_record['Year'].tolist()[0]),
                                    str(df_record['Month'].tolist()[0]), str(df_record['Project'].tolist()[0]),
                                    str(df_record['Audit type'].tolist()[0]), score, score_range]
    return (result)


def format_modification(df):
    # print(df)
    df = df[['Vendor', 'Auditor', 'By Audit Category', 'Project', 'Product Category', 'Year', 'Month', 'Week', 'Line',
             'Process', 'Issue category', 'Sub category', 'Severity', 'Status', 'Findings',
             'Failure Analysis/Root cause', 'Corrective action']]



    # print(df.columns.values.tolist)
    new_name_list = ['vendor', 'Auditor', 'Audit type', 'Project', 'Compnent', 'Year', 'Month', 'Week', 'Line',
                     'Station', 'Issue category', 'Sub-category', 'Issue sevrity', 'MIL Status', 'MIL Description ',
                     'FA', 'CA']
    column_name = {}
    ori_name = df.columns.values.tolist()
    for i in range(len(df.columns.values.tolist())):
        column_name[ori_name[i]] = new_name_list[i]

    # print(column_name)

    df.rename(columns=column_name, inplace=True)

    df['QCP/DFM requirement'] = [''] * df.shape[0]
    df['Actual status'] = [''] * df.shape[0]
    df['FACA reject apple comments'] = [''] * df.shape[0]
    df['Repeat issue(Y/N)'] = [''] * df.shape[0]
    df['FACA response(Y/N)'] = [''] * df.shape[0]
    df['Block parts Q\'ty'] = [''] * df.shape[0]
    df['Scrap parts Q\'ty'] = [''] * df.shape[0]
    df['Related Program'] = [''] * df.shape[0]
    df['Error proof actions'] = [''] * df.shape[0]
    df['Highlight Y/N'] = [''] * df.shape[0]
    df['Highlight Description'] = [''] * df.shape[0]
    df['123'] = [''] * df.shape[0]

    return (df)


def to_datebase_modify(data):
    for i in data.columns:
        data[i] = data[i].fillna('')
        return(data)


# excel表格转json文件
def excel_to_json(excel_file):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    # print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            if cell.value == None:
                value = ''
            else:
                value = cell.value
            one_line[k] = value
        # print(one_line)
        result.append(one_line)
    book.close()
    return result


def upload_mil_to_tableau(request):
    if request.method == 'GET':
        engine = create_engine(
            'postgresql+psycopg2://smartaudit:%s@isdsepgdb-mqa-smartaudit-prod.g.apple.com:5301/smartaudit_prod' % urllib.parse.quote_plus(
                '6AlDS:5~RC,Q0A-G@i5s9(Q+!ijzAK'))
        engine.dispose()
        return HttpResponse('OK')
    else:

        userfrom = UserForm(request.POST, request.FILES)
        if userfrom.is_valid():
            filename = userfrom.cleaned_data['filename']
            json_list = excel_to_json(userfrom.cleaned_data['filename'])
            temp_json1 = json.dumps(json_list, ensure_ascii=False)
            temp_json = json.loads(temp_json1)
            # temp_json = json.loads(request.POST.get("file"))
            temp_df = pd.DataFrame()
            for i in temp_json:
                temp_df = pd.concat([temp_df, pd.DataFrame([i])])

            temp_df.reset_index(inplace=True)

            temp_database = format_modification(temp_df)
            # temp_database.to_excel('a.xlsx')
            # print(temp_database)
            temp_project_score = get_by_project_score(temp_database)
            # temp_project_score.to_excel('b.xlsx')
            # print(temp_project_score)
            #pwd = ''
            engine = create_engine('postgresql+psycopg2://smartaudit:%s@isdsepgdb-mqa-smartaudit-prod.g.apple.com:5301/smartaudit_prod' % urllib.parse.quote_plus('6AlDS:5~RC,Q0A-G@i5s9(Q+!ijzAK'))
            temp_database = to_datebase_modify(temp_database)
            temp_project_score = to_datebase_modify(temp_project_score)
            temp_database.to_sql(name='test_T_database', con=engine, if_exists="append", index=False)
            temp_project_score.to_sql(name='test_T_project_score', con=engine, if_exists="append", index=False)
            # print("finished")
            engine.dispose()
            return HttpResponse('upload ok')